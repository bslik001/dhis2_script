#!/usr/bin/env python3
"""
pivot_tracked_and_stage.py

SCRIPT ORCHESTRATEUR COMPLET

Ce script regroupe désormais **3 grandes étapes**, sans modifier leur logique interne :

1️⃣ Téléchargement des Tracked Entity Instances (ex-download_tracked.py)
2️⃣ Téléchargement des Events / Program Stages (ex-download.py)
3️⃣ Pivot + génération du fichier Excel (code existant)

⚠️ IMPORTANT :
- Aucune logique n’a été modifiée
- Les fonctions ont simplement été regroupées
- Les paramètres restent configurés via CLI > .env > défauts
"""

# ==========================================================
# ======================= IMPORTS ==========================
# ==========================================================

import argparse
import json
import math
import os
import sys
from pathlib import Path
from urllib.parse import urlencode
import re

import pandas as pd
import requests
from dotenv import load_dotenv
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ==========================================================
# ================== FONCTIONS COMMUNES ===================
# ==========================================================

def build_url(base_url, params):
    """
    Construit une URL complète DHIS2 à partir :
    - de l’URL de base
    - d’un dictionnaire de paramètres
    """
    query_string = urlencode(params)
    return f"{base_url}?{query_string}" if query_string else base_url


# ==========================================================
# ========== 1️⃣ DOWNLOAD TRACKED ENTITY INSTANCES =========
# ==========================================================

def download_dhis2_tracked(base_url, params, output_file, token):
    """
    Télécharge le CSV des Tracked Entity Instances depuis DHIS2
    (code IDENTIQUE à download_tracked.py)
    """
    full_url = build_url(base_url, params)
    print(f"URL générée (TRACKED) : {full_url}")

    headers = {"Authorization": f"ApiToken {token}"}

    try:
        response = requests.get(full_url, headers=headers, stream=True, timeout=60)
        response.raise_for_status()

        total_size = int(response.headers.get("content-length", 0))

        with open(output_file, "wb") as f, tqdm(
            total=total_size,
            unit="B",
            unit_scale=True,
            unit_divisor=1024,
            desc="Téléchargement Tracked"
        ) as bar:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
                    bar.update(len(chunk))

        print(f"✅ Tracked Entity Instances téléchargés : {output_file}")

    except requests.exceptions.RequestException as e:
        print(f"❌ Erreur téléchargement tracked : {e}")
        sys.exit(1)


# ==========================================================
# ========== 2️⃣ DOWNLOAD EVENTS / PROGRAM STAGES ==========
# ==========================================================

def download_dhis2_events(base_url, params, output_file, token):
    """
    Télécharge le CSV des Events (Program Stages)
    (code IDENTIQUE à download.py)
    """
    full_url = build_url(base_url, params)
    print(f"URL générée (EVENTS) : {full_url}")

    headers = {"Authorization": f"ApiToken {token}"}

    try:
        response = requests.get(full_url, headers=headers, stream=True, timeout=60)
        response.raise_for_status()

        total_size = int(response.headers.get("content-length", 0))

        with open(output_file, "wb") as f, tqdm(
            total=total_size,
            unit="B",
            unit_scale=True,
            unit_divisor=1024,
            desc="Téléchargement Events"
        ) as bar:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
                    bar.update(len(chunk))

        print(f"✅ Events téléchargés : {output_file}")

    except requests.exceptions.RequestException as e:
        print(f"❌ Erreur téléchargement events : {e}")
        sys.exit(1)


# ==========================================================
# ======================= ÉTAT =============================
# ==========================================================

def load_state(state_file: str) -> dict:
    if os.path.exists(state_file):
        with open(state_file, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"completed_stages": []}

def save_state(state_file: str, state: dict):
    Path(state_file).parent.mkdir(parents=True, exist_ok=True)
    with open(state_file, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)

def clear_state(state_file: str):
    if os.path.exists(state_file):
        os.remove(state_file)

# ==========================================================
# ======================= API DHIS2 ========================
# ==========================================================

def safe_get(url: str, token: str, params: dict | None = None) -> dict:
    headers = {"Authorization": f"ApiToken {token}"}
    response = requests.get(url, headers=headers, params=params, timeout=60)
    response.raise_for_status()
    return response.json()

def get_program_stages(base_url: str, token: str, program_id: str) -> list[dict]:
    url = f"{base_url}/programs/{program_id}/programStages.json"
    params = {"fields": "id,displayName,sortOrder", "paging": "false"}
    data = safe_get(url, token, params)
    return sorted(data.get("programStages", []), key=lambda s: s.get("sortOrder", math.inf))

def get_stage_dataelements(base_url: str, token: str, stage_id: str) -> list[str]:
    url = f"{base_url}/programStages/{stage_id}.json"
    params = {
        "fields": "programStageDataElements[dataElement[id,displayName],sortOrder]",
        "paging": "false"
    }
    data = safe_get(url, token, params)
    elements = sorted(
        data.get("programStageDataElements", []),
        key=lambda x: x.get("sortOrder", math.inf)
    )
    return [e["dataElement"]["displayName"] for e in elements if e.get("dataElement")]

# ==========================================================
# ===================== MAPPING =============================
# ==========================================================

def load_mapping(mapping_file: str) -> dict:
    if os.path.exists(mapping_file):
        with open(mapping_file, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_mapping(mapping_file: str, mapping: dict):
    Path(mapping_file).parent.mkdir(parents=True, exist_ok=True)
    with open(mapping_file, "w", encoding="utf-8") as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)

def build_de_mapping_from_api(base_url: str, token: str) -> dict:
    url = f"{base_url}/dataElements.json"
    params = {"paging": "false", "fields": "id,displayName"}
    data = safe_get(url, token, params)
    return {de["id"]: de["displayName"] for de in data.get("dataElements", [])}

# ==========================================================
# ================= MAP VALUES (from map_values.py) ========
# ==========================================================
def charger_fichier(path: str) -> pd.DataFrame:
    """Charge un fichier .csv ou .xlsx en DataFrame (dtype=str).
    Similaire à la fonction dans map_values.py mais limitée au besoin présent.
    """
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return pd.read_csv(path, dtype=str, keep_default_na=False)
    elif ext == ".xlsx" or ext == ".xls":
        return pd.read_excel(path, dtype=str)
    else:
        raise ValueError(f"Format non supporté : {ext}. Utilisez .csv ou .xlsx")


def nettoyer_valeur(val):
    """Nettoie une valeur selon les règles utilisées par map_values.py :
    - trim
    - true/false -> '1'/'0'
    - dates 'YYYY-MM-DD 00:00:00.0' -> 'YYYY-MM-DD'
    - floats style '19.0' -> '19'
    """
    if not isinstance(val, str):
        return val
    val = val.strip()

    # if val.lower() == "true":
    #     return "1"
    # if val.lower() == "false":
    #     return "0"

    if re.match(r"^\d{4}-\d{2}-\d{2}(?: 00:00:00(?:\.0)?)?$", val):
        return val.split(" ")[0]

    if re.match(r"^\d+\.0$", val):
        try:
            return str(int(float(val)))
        except Exception:
            return val

    return val


def nettoyer_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    # Appliquer nettoyage cellule par cellule
    return df.apply(
        lambda col: col.map(
            lambda x: nettoyer_valeur(x) if pd.notna(x) else x
        )
    )


def construire_mapping(correspondance_df: pd.DataFrame, structure_df: pd.DataFrame,
                       col_a1: str, col_a2: str, col_b1: str, col_b2: str) -> dict:
    """Construire le mapping par colonne à partir des deux fichiers (même logique que map_values.py).
    Retourne un dict: {target_column_name: {value_in_source: replacement_code}}
    """
    mapping_par_colonne = {}

    # Normaliser noms de colonnes
    correspondance_df = correspondance_df.copy()
    structure_df = structure_df.copy()
    correspondance_df.columns = correspondance_df.columns.str.strip()
    structure_df.columns = structure_df.columns.str.strip()

    for _, struct in structure_df.iterrows():
        b1_val = str(struct.get(col_b1, "")).strip()
        b2_col = str(struct.get(col_b2, "")).strip()

        if not b2_col:
            continue

        for _, row in correspondance_df.iterrows():
            a1_val = str(row.get(col_a1, "")).strip()
            a2_val = str(row.get(col_a2, "")).strip()

            if b1_val != a1_val:
                continue

            valeurs = [v.strip() for v in a1_val.split(";")] if a1_val else []
            codes = [c.strip().replace(":", "") for c in a2_val.split(";")] if a2_val else []

            if len(valeurs) != len(codes) or not valeurs:
                continue

            mapping_par_colonne[b2_col] = dict(zip(valeurs, codes))
            break

    return mapping_par_colonne


def appliquer_mapping_a_df(df: pd.DataFrame, mapping_par_colonne: dict):
    """Applique le mapping sur un DataFrame et retourne (df_modifie, log, lignes_modifiees)
    similaire à la sortie de map_values.py.
    """
    df = df.copy()
    log = []
    lignes_modifiees = set()

    for col in df.columns:
        if col in mapping_par_colonne:
            mapping = mapping_par_colonne[col]
            for i, val in df[col].items():
                val_clean = str(val).strip() if pd.notna(val) else ""
                if val_clean in mapping:
                    nouveau = mapping[val_clean]
                    log.append(f"Ligne {i}, onglet-col '{col}' : '{val_clean}' -> '{nouveau}'")
                    df.at[i, col] = nouveau
                    lignes_modifiees.add(i)

    return df, log, lignes_modifiees


def apply_mapping_to_excel(correspondance_path: str,
                           structure_path: str,
                           excel_path: str,
                           col_a1: str = "Option Codes",
                           col_a2: str = "Option Details",
                           col_b1: str = "Option codes",
                           col_b2: str = "DataElement",
                           log_file: str | None = None,
                           preview: bool = False):
    """Applique la correspondance (comme map_values.py) à CHAQUE onglet de `excel_path`.

    La fonction lit `correspondance_path` et `structure_path` (csv/xlsx), construit le mapping
    puis itère sur tous les onglets du fichier Excel `excel_path`, applique les remplacements
    et ré-écrit le fichier (écrasement).
    """
    # Chargement des fichiers de correspondance/structure
    corr_df = charger_fichier(correspondance_path)
    struct_df = charger_fichier(structure_path)

    corr_df.columns = corr_df.columns.str.strip()
    struct_df.columns = struct_df.columns.str.strip()

    mapping_par_col = construire_mapping(corr_df, struct_df, col_a1, col_a2, col_b1, col_b2)

    if not mapping_par_col:
        print("⚠️ Aucun mapping trouvé : vérifiez les fichiers de correspondance/structure et les noms de colonnes.")

    # Lire tous les onglets
    all_sheets = pd.read_excel(excel_path, sheet_name=None, dtype=str)

    total_logs = []

    for sheet_name, df in all_sheets.items():
        print(f"▶️ Application mapping sur onglet : {sheet_name}")
        df = df.fillna("")
        df = nettoyer_dataframe(df)
        new_df, logs, _ = appliquer_mapping_a_df(df, mapping_par_col)
        all_sheets[sheet_name] = new_df
        total_logs.extend([f"[{sheet_name}] {l}" for l in logs])
        print(f"  - Remplacements appliqués dans '{sheet_name}' : {len(logs)}")

    # Écrire l'Excel en écrasant
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        for sheet_name, df in all_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    if log_file:
        with open(log_file, "w", encoding="utf-8") as f:
            for ligne in total_logs:
                f.write(ligne + "\n")
        print(f"🧾 Journal enregistré sous : {log_file}")
    # else:
    #     # print compact
    #     for ligne in total_logs:
    #         print("  -", ligne)

    if preview:
        print("\n🔍 Aperçu des premiers changements :")
        for sheet_name, df in all_sheets.items():
            print(f"-- Onglet: {sheet_name}")
            print(df.head(5).to_string(index=False))


# ==========================================================
# ======================= PIVOTS ============================
# ==========================================================

def pivot_tracked_df(input_file: str, aggfunc: str) -> pd.DataFrame:
    df = pd.read_csv(input_file, dtype=str)
    pivot = df.pivot_table(
        index="trackedEntityInstance",
        columns="displayName",
        values="value",
        aggfunc=aggfunc
    ).reset_index().fillna("")
    pivot.insert(
        pivot.columns.get_loc("trackedEntityInstance") + 1,
        "ID",
        range(1, len(pivot) + 1)
    )
    return pivot

def run_pivot_and_excel(
    tracked_input: str,
    stage_input: str,
    output: str,
    base_url: str,
    token: str,
    program: str,
    aggfunc: str,
    mapping_file: str,
    state_file: str,
    strict: bool,
):
    """
    Exécute TOUTE la logique existante de pivot et génération Excel.
    ⚠️ Code IDENTIQUE à l’existant, simplement encapsulé.
    """

    # ======================
    # Chargement de l’état
    # ======================
    state = load_state(state_file)

    # ======================
    # Pivot Tracked Entities
    # ======================
    tracked_df = pivot_tracked_df(tracked_input, aggfunc)
    print(f"▶️ Pivot du premier onglet : {len(tracked_df)} lignes")

    cols = list(tracked_df.columns)
    serial_cols = [c for c in cols if "_serial_number" in str(c).lower()]
    date_cols = [c for c in cols if "_date_" in str(c).lower()]
    parent_cols = [c for c in cols if "_parent_consent" in str(c).lower()]

    if strict:
        ordered_cols = ["trackedEntityInstance"] + serial_cols + ["ID"] + date_cols + parent_cols
    else:
        ordered_cols = ["trackedEntityInstance"] + serial_cols + ["ID"]
        remaining = [c for c in cols if c not in ordered_cols]
        ordered_cols += sorted(remaining)

    tracked_df = tracked_df[ordered_cols]

    # ======================
    # Écriture premier onglet
    # ======================
    if os.path.exists(output) and not os.path.exists(state_file):
        wb = load_workbook(output)
        if "TrackedEntities" not in wb.sheetnames:
            writer = pd.ExcelWriter(output, engine="openpyxl", mode="a", if_sheet_exists="overlay")
            write_with_progress(tracked_df, writer, "TrackedEntities")
            writer.close()
        else:
            print("⏩ Onglet TrackedEntities déjà présent")
    else:
        writer = pd.ExcelWriter(output, engine="openpyxl")
        write_with_progress(tracked_df, writer, "TrackedEntities")
        writer.close()

    # ======================
    # Réouverture pour stages
    # ======================
    writer = pd.ExcelWriter(output, engine="openpyxl", mode="a", if_sheet_exists="overlay")

    df = pd.read_csv(stage_input, dtype=str)

    # ======================
    # Mapping dataElements
    # ======================
    mapping = load_mapping(mapping_file)
    missing = [u for u in df["dataElement"].unique() if u not in mapping]

    if missing:
        api_map = build_de_mapping_from_api(base_url, token)
        for uid in missing:
            mapping[uid] = api_map.get(uid, uid)
        save_mapping(mapping_file, mapping)

    df["dataElement"] = df["dataElement"].map(mapping)

    # ======================
    # Pivot global events
    # ======================
    pivot = df.pivot_table(
        index="enrollment",
        columns="dataElement",
        values="value",
        aggfunc=aggfunc
    ).reset_index().fillna("")

    pivot.insert(
        pivot.columns.get_loc("enrollment") + 1,
        "ID",
        range(1, len(pivot) + 1)
    )

    stages = get_program_stages(base_url, token, program)

    try:
        for stage in stages:
            stage_name = stage["displayName"][:31]

            if stage_name in state["completed_stages"]:
                print(f"⏩ Déjà traité : {stage_name}")
                continue


            print(f"▶️ Traitement : {stage_name}")

            # Récupère la liste ordonnée des dataElements pour ce stage (displayName)
            elements = get_stage_dataelements(base_url, token, stage["id"])  # list[str]

            # Construire les colonnes du sheet dans l'ordre exact donné par le ProgramStage
            cols = ["enrollment", "ID"] + elements

            # Reindexer le pivot global pour garantir que toutes les colonnes présentes dans
            # la définition du stage sont écrites, dans l'ordre demandé. Les colonnes
            # manquantes (pas présentes dans les events) seront créées et remplies vides.
            # Cela évite d'omettre un ProgramStage simplement parce que toutes ses valeurs
            # seraient vides : la feuille contiendra alors les colonnes (vides) correspondantes.
            sheet_df = pivot.reindex(columns=cols, fill_value="").copy()
            sheet_df = sheet_df.fillna("")

            # Détecter si toutes les colonnes de dataElement sont vides pour tous les enrollments
            data_columns = [c for c in cols if c not in ("enrollment", "ID")]
            non_empty = False
            if data_columns:
                # any non-empty cell across the data columns?
                non_empty = sheet_df[data_columns].apply(lambda col: col.astype(str).str.strip().ne("")).any().any()

            if not non_empty:
                print("⚠️ Feuille vide : toutes les valeurs sont vides")

            # Écrire la feuille (même si toutes les colonnes de dataElement sont vides)
            sheet_df.to_excel(writer, sheet_name=stage_name, index=False)
            auto_adjust_column_width(writer.sheets[stage_name])

            state["completed_stages"].append(stage_name)
            save_state(state_file, state)

            print(f"✅ Terminé : {stage_name}")

    finally:
        writer.close()

    clear_state(state_file)
    print("\n🎉 Fichier Excel généré avec succès")

# ==========================================================
# ===================== EXCEL UTIL =========================
# ==========================================================

def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            value = str(cell.value) if cell.value else ""
            max_length = max(max_length, len(value))
        ws.column_dimensions[column].width = max_length + 2

def write_with_progress(df: pd.DataFrame, writer, sheet_name: str, chunk_size: int = 50):
    total = len(df)
    df.head(0).to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]

    for start in range(0, total, chunk_size):
        end = min(start + chunk_size, total)
        chunk = df.iloc[start:end]
        for r_idx, row in enumerate(chunk.values, start=start + 2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        progress = (end / total) * 100
        sys.stdout.write(f"\r   → Lignes {start+1} à {end} / {total} ({progress:.1f}%)")
        sys.stdout.flush()

    auto_adjust_column_width(ws)
    print(f"\n✅ Onglet '{sheet_name}' écrit avec succès\n")

# ==========================================================
# ========================= MAIN ============================
# ==========================================================

def main():
    load_dotenv()

    parser = argparse.ArgumentParser(description="Pipeline DHIS2 complet")
    parser.add_argument("--skip-download", action="store_true")
    parser.add_argument("--only-download", action="store_true")
    parser.add_argument("--only-pivot", action="store_true")
    parser.add_argument("--apply-mapping", action="store_true", help="Appliquer le mapping (data/correspondance.csv & data/structure.xlsx) sur l'Excel généré")
    parser.add_argument("--mapping-correspondance", default="data/correspondance.csv", help="Chemin vers le fichier de correspondance (csv/xlsx)")
    parser.add_argument("--mapping-structure", default="data/structure.xlsx", help="Chemin vers le fichier structure (csv/xlsx)")
    parser.add_argument("--mapping-col-a1", default="Option Codes", help="Nom colonne A1 dans correspondance")
    parser.add_argument("--mapping-col-a2", default="Option Details", help="Nom colonne A2 dans correspondance")
    parser.add_argument("--mapping-col-b1", default="Option codes", help="Nom colonne B1 dans structure")
    parser.add_argument("--mapping-col-b2", default="DataElement", help="Nom colonne B2 dans structure qui contient le nom de colonne cible")
    parser.add_argument("--mapping-log-file", default=None, help="Fichier log pour les remplacements (optionnel)")
    parser.add_argument("--mapping-preview", action="store_true", help="Afficher un aperçu après application du mapping")

    args, _ = parser.parse_known_args()

    # Sécurité : flags incompatibles
    if args.only_download and args.only_pivot:
        parser.error("--only-download et --only-pivot sont incompatibles")

    # ======================================================
    # DÉTERMINATION DU MODE D’EXÉCUTION
    # ======================================================

    do_download = True
    do_pivot = True

    if args.only_download:
        do_pivot = False

    if args.only_pivot:
        do_download = False

    if args.skip_download:
        do_download = False

    # ======================================================
    # ÉTAPE 1 : DOWNLOAD
    # ======================================================
    if do_download:
        print("\n📥 ÉTAPE : Téléchargement DHIS2\n")

        download_dhis2_tracked(
            base_url=os.getenv("TRACKED_BASE_URL"),
            params={
                "program": os.getenv("TRACKED_PROGRAM"),
                "programStartDate": os.getenv("TRACKED_PROGRAM_START_DATE"),
                "programEndDate": os.getenv("TRACKED_PROGRAM_END_DATE"),
                "ouMode": os.getenv("TRACKED_OU_MODE"),
                "format": os.getenv("TRACKED_FORMAT"),
            },
            output_file=os.getenv("TRACKED_OUTPUT"),
            token=os.getenv("PIVOT_TOKEN"),
        )

        download_dhis2_events(
            base_url=os.getenv("DOWNLOAD_BASE_URL"),
            params={
                "orgUnit": os.getenv("DOWNLOAD_ORG_UNIT"),
                "program": os.getenv("DOWNLOAD_PROGRAM"),
                "startDate": os.getenv("DOWNLOAD_START_DATE"),
                "endDate": os.getenv("DOWNLOAD_END_DATE"),
                "ouMode": os.getenv("DOWNLOAD_OU_MODE"),
                "skipPaging": os.getenv("DOWNLOAD_SKIP_PAGING"),
                "format": os.getenv("DOWNLOAD_FORMAT"),
            },
            output_file=os.getenv("PIVOT_INPUT"),
            token=os.getenv("PIVOT_TOKEN"),
        )
    else:
        print("⏩ Téléchargement ignoré")

    # ======================================================
    # ÉTAPE 2 : PIVOT
    # ======================================================
    if do_pivot:
        print("\n📊 ÉTAPE : Pivot & Excel\n")

        run_pivot_and_excel(
            tracked_input=os.getenv("TRACKED_OUTPUT"),
            stage_input=os.getenv("PIVOT_INPUT"),
            output=os.getenv("MERGED_PIVOT_OUTPUT"),
            base_url=os.getenv("PIVOT_BASE_URL"),
            token=os.getenv("PIVOT_TOKEN"),
            program=os.getenv("DOWNLOAD_PROGRAM"),
            aggfunc=os.getenv("PIVOT_AGGFUNC", "first"),
            mapping_file=os.getenv("PIVOT_MAPPING_FILE"),
            state_file=os.getenv("PIVOT_STATE_FILE"),
            strict=False,
        )

        # Optionnel : appliquer le mapping sur l'Excel généré
        if args.apply_mapping:
            excel_file = os.getenv("MERGED_PIVOT_OUTPUT")
            mapping_corr = args.mapping_correspondance
            mapping_struct = args.mapping_structure
            if excel_file and os.path.exists(excel_file):
                print(f"\n🧭 Application du mapping sur {excel_file}")
                apply_mapping_to_excel(
                    correspondance_path=mapping_corr,
                    structure_path=mapping_struct,
                    excel_path=excel_file,
                    col_a1=args.mapping_col_a1,
                    col_a2=args.mapping_col_a2,
                    col_b1=args.mapping_col_b1,
                    col_b2=args.mapping_col_b2,
                    log_file=args.mapping_log_file,
                    preview=args.mapping_preview,
                )
            else:
                print(f"⚠️ Fichier Excel introuvable pour mapping: {excel_file}")
    else:
        print("⏩ Pivot ignoré")

    print("\n✅ Pipeline terminé")


if __name__ == "__main__":
    main()
